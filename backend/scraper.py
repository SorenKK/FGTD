from bs4 import BeautifulSoup
from selenium import webdriver
#from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver import ActionChains
import re
import pandas as pd
from Bio import Entrez,Medline
from io import StringIO
import random
import string
import os
import sys
import functools
import logging
import io
import urllib.parse
import json
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font
from selenium.common.exceptions import UnexpectedAlertPresentException
from math import log1p
from datetime import datetime, date
from datetime import datetime, date
# --- NUOVI IMPORT PER QDRANT ---
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct
from sentence_transformers import SentenceTransformer
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[logging.StreamHandler(sys.stdout)])
# --- INIZIALIZZAZIONE AI ---
logging.info("Loading embedding model (all-MiniLM-L6-v2)...")
# Scarica il modello (80MB) al primo avvio, poi usa la cache.
# Se non trova GPU, usa automaticamente la CPU.
model = SentenceTransformer('all-MiniLM-L6-v2') 

# Inizializza Qdrant in memoria (RAM). I dati spariscono alla chiusura, perfetto per questo uso.
qdrant = QdrantClient(":memory:")

####################################################FUNZIONI PER I FILTRI
def _expand_filters(driver):
    try:
        driver.find_element(By.ID, "more_filter_groups_link").click()
        time.sleep(0.5)
        for gid in ["fg_subVarTypeGds", "fg_sampleCountGds",
                    "fg_suppFileGds", "fg_field_search"]:
            cb = driver.find_element(By.ID, gid)
            if not cb.is_selected():
                cb.click()
        driver.find_element(By.ID, "filter_groups_apply").click()
        time.sleep(0.3)
    except Exception as e:
        logging.warning(f"Expand filters failed: {e}")

def autocomplete_filter(driver, filter_id, values, apply_btn_id):
    wait = WebDriverWait(driver, 4)
    values = [v.strip() for v in values if v.strip()]
    if not values:
        logging.debug(f"[{filter_id}] No values")
        return

    group_label = filter_id[:-3].capitalize()
    for v in values:
        logging.debug(f"[{filter_id}] Apro popup per «{v}»")
        wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, f"a[href='#{filter_id}_more']"))
        ).click()

        inp = wait.until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, "div.facets_dialog input.of_sel_inp")))
        driver.execute_script("arguments[0].value = '';", inp)
        inp.send_keys(v)
        wait.until(lambda d: inp.get_attribute("value").strip() != "")
        inp.send_keys(Keys.RETURN)


        try:
            sidebar_ul = driver.find_element(
                By.XPATH,
                f"//li[contains(@class,'filter_grp') and .//h3[normalize-space()='{group_label}']]//ul"
            )
            link = sidebar_ul.find_element(
                By.CSS_SELECTOR,
                f"a[data-value_id='{v.lower()}']"
            )
            ActionChains(driver).move_to_element(link).click().perform()
            WebDriverWait(driver, 4).until(EC.staleness_of(link))  # opzionale
        except Exception as e:
            logging.warning(f"[{filter_id}] : {e}")

    try:
        btn = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.ID, apply_btn_id)))
        driver.execute_script("""
            arguments[0].scrollIntoView({block:'center'});
            arguments[0].click();
        """, btn)
        WebDriverWait(driver, 5).until(EC.invisibility_of_element_located((By.ID, apply_btn_id)))
    except TimeoutException:
        logging.info(f"[{filter_id}] Apply ({apply_btn_id}) continue")


def apply_study_type_filter(driver, labels):
    time.sleep(0.5)
    wait = WebDriverWait(driver, 4)

    # 1) Attendi che qualsiasi dialog precedente sia sparito
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.facets_dialog")))

    # 2) Trova e clicca il link Customize… di Study type
    customize = wait.until(EC.element_to_be_clickable((
        By.XPATH, "//a[@href='#studyTypeGds_more' and contains(.,'Customize')]"
    )))
    ActionChains(driver).move_to_element(customize).click().perform()
    logging.debug("[Study type] Customize…")
    #time.sleep(0.3)

    # 3) Attendi il popup e prendi l'HTML
    popup = wait.until(EC.visibility_of_element_located((
        By.CSS_SELECTOR, "#studyTypeGds_more .facets_dialog"
    )))
    html = popup.get_attribute("innerHTML")
    soup = BeautifulSoup(html, "html.parser")

    # 4) Costruisci mapping label → data-value_id
    all_items = soup.select("ul.facet_more[data-filter_id='studyTypeGds'] li")
    mapping = {
        li.find("label").get_text(strip=True): li.find("input")["data-value_id"]
        for li in all_items if li.find("input") and li.find("label")
    }

    # 5) Deseleziona tutti i checkbox selezionati
    for li in all_items:
        chk = li.find("input")
        if chk and chk.has_attr("checked"):
            dv = chk["data-value_id"]
            try:
                cb = wait.until(EC.element_to_be_clickable((
                    By.CSS_SELECTOR, f"#studyTypeGds_more input[data-value_id='{dv}']"
                )))
                if cb.is_selected():
                    cb.click()
            except Exception as e:
                logging.warning(f"[Study type] {dv}: {e}")

    # 6) Seleziona le checkbox richieste
    selected_dv_ids = []
    for label in labels:
        dv = mapping.get(label)
        if not dv:
            logging.warning(f"[Study type] `{label}` not found")
            continue
        chk = wait.until(EC.element_to_be_clickable((
            By.CSS_SELECTOR, f"#studyTypeGds_more input[data-value_id='{dv}']"
        )))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", chk)
        if not chk.is_selected():
            chk.click()
            logging.debug(f"[Study type] Selection of `{label}`")
        selected_dv_ids.append(dv)
        #time.sleep(0.3)

    # 7) Click su Show e attendi chiusura del dialog
    show_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#studyTypeGds_apply")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", show_btn)
    driver.execute_script("arguments[0].click();", show_btn)
    logging.debug("[Study type] Show")
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "#studyTypeGds_more .facets_dialog")))
    time.sleep(0.2)

    # 8) Clicca nella sidebar le stesse voci
    group_label = "Study type"
    for dv in selected_dv_ids:
        try:
            sidebar_ul = driver.find_element(
                By.XPATH,
                f"//li[contains(@class,'filter_grp') and .//h3[normalize-space()='{group_label}']]//ul"
            )
            link = sidebar_ul.find_element(
                By.CSS_SELECTOR,
                f"a[data-value_id='{dv}']"
            )
            logging.debug(f"[Study type] found «{dv}» in sidebar")
            ActionChains(driver).move_to_element(link).click().perform()
            time.sleep(0.5)
        except Exception as e:
            logging.warning(f"[Study type] «{dv}»: {e}")

def click_filter_links(driver, group_label,labels):
    time.sleep(0.5)
    logging.info("[Subset variable type]")
    wait = WebDriverWait(driver, 4)

    # 1) Attendi che qualsiasi dialog precedente sia sparito
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.facets_dialog")))

    # 2) Trova e clicca il link Customize… di Subset variable type
    customize = wait.until(EC.element_to_be_clickable((
        By.XPATH, "//a[@href='#subVarTypeGds_more' and contains(.,'Customize')]"
    )))
    ActionChains(driver).move_to_element(customize).click().perform()
    logging.debug("[Subset variable] Customize…")
    time.sleep(0.2)

    # 3) Attendi il popup e prendi l'HTML
    popup = wait.until(EC.visibility_of_element_located((
        By.CSS_SELECTOR, "#subVarTypeGds_more .facets_dialog"
    )))
    html = popup.get_attribute("innerHTML")
    soup = BeautifulSoup(html, "html.parser")

    # 4) Costruisci mapping label → data-value_id
    all_items = soup.select("ul.facet_more[data-filter_id='subVarTypeGds'] li")
    mapping = {
        li.find("label").get_text(strip=True): li.find("input")["data-value_id"]
        for li in all_items if li.find("input") and li.find("label")
    }

    # 5) Deseleziona tutti i checkbox selezionati
    for li in all_items:
        chk = li.find("input")
        if chk and chk.has_attr("checked"):
            dv = chk["data-value_id"]
            try:
                cb = wait.until(EC.element_to_be_clickable((
                    By.CSS_SELECTOR, f"#subVarTypeGds_more input[data-value_id='{dv}']"
                )))
                if cb.is_selected():
                    cb.click()
                    logging.debug(f"[Subset variable] Deselezionato {dv}")
            except Exception as e:
                logging.warning(f"[Subset variable] Impossibile deselezionare {dv}: {e}")

    # 6) Seleziona le checkbox richieste
    selected_dv_ids = []
    for label in labels:
        dv = mapping.get(label)
        if not dv:
            logging.warning(f"[Subset variable] Voce `{label}` non trovata")
            continue
        chk = wait.until(EC.element_to_be_clickable((
            By.CSS_SELECTOR, f"#subVarTypeGds_more input[data-value_id='{dv}']"
        )))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", chk)
        if not chk.is_selected():
            chk.click()
            logging.debug(f"[Subset variable] Selezionato `{label}`")
        selected_dv_ids.append(dv)
        #time.sleep(0.2)

    # 7) Click su Show e attendi chiusura del dialog
    show_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#subVarTypeGds_apply")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", show_btn)
    driver.execute_script("arguments[0].click();", show_btn)
    logging.debug("[Subset variable]")
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "#subVarTypeGds_more .facets_dialog")))
    time.sleep(0.2)

    # 8) Clicca nella sidebar le stesse voci
    group_label = "Subset variable type"
    for dv in selected_dv_ids:
        try:
            sidebar_ul = driver.find_element(
                By.XPATH,
                f"//li[contains(@class,'filter_grp') and .//h3[normalize-space()='{group_label}']]//ul"
            )
            link = sidebar_ul.find_element(
                By.CSS_SELECTOR,
                f"a[data-value_id='{dv}']"
            )
            logging.debug(f"[Subset variable] «{dv}» in sidebar")
            ActionChains(driver).move_to_element(link).click().perform()
            time.sleep(0.3)
        except Exception as e:
            logging.warning(f"[Subset variable] Non ho potuto cliccare «{dv}»: {e}")

    logging.info("finished [Subset variable type]")

def apply_publication_date_range(driver, start_date, end_date):
    time.sleep(0.5)
    wait = WebDriverWait(driver, 4)

    # Apri il popup
    driver.find_element(By.ID, "facet_date_rangepubDatesGds").click()

    def set_field(field_id, val):
        logging.info(f"📝 Set {field_id} = {val}")
        field = wait.until(EC.presence_of_element_located((By.ID, field_id)))
        field.click()
        field.clear()
        field.send_keys(str(val))

    # START DATE
    set_field("facet_date_st_yearpubDatesGds", f"{start_date[0]:04d}")   # YYYY
    set_field("facet_date_st_monthpubDatesGds", f"{start_date[1]:02d}")  # MM
    set_field("facet_date_st_daypubDatesGds", f"{start_date[2]:02d}")    # DD

    # END DATE
    set_field("facet_date_end_yearpubDatesGds", f"{end_date[0]:04d}")    # YYYY
    set_field("facet_date_end_monthpubDatesGds", f"{end_date[1]:02d}")   # MM
    set_field("facet_date_end_daypubDatesGds", f"{end_date[2]:02d}")     # DD

    # Clicca su Apply
    time.sleep(0.3)
    apply_btn = wait.until(EC.element_to_be_clickable((By.ID, "facet_date_range_applypubDatesGds")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", apply_btn)
    driver.execute_script("arguments[0].click();", apply_btn)
    logging.info("[Date range] Filter applied")

def apply_supplementary_file_filter(driver, values):
    logging.info("Supplementary file")
    wait = WebDriverWait(driver, 4)

    values = [v.strip().upper() for v in values if v.strip()]
    print(values)
    if not values:
        logging.info("No supplementary files.")
        return

    default_ids = ["celGds", "gprGds", "wigGds", "bedGds"]
    added_ids = []

    for idx, val in enumerate(values):
        # Riapre sempre il Customize...
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.facets_dialog")))
        customize = wait.until(EC.element_to_be_clickable((
            By.XPATH, "//a[@href='#suppFileGds_more' and contains(.,'Customize')]"
        )))
        ActionChains(driver).move_to_element(customize).click().perform()
        logging.info("[Supplementary file] Customize...")
        time.sleep(0.3)

        # Attendi popup e prendi HTML
        popup = wait.until(EC.visibility_of_element_located((
            By.CSS_SELECTOR, "#suppFileGds_more .facets_dialog"
        )))
        html = popup.get_attribute("innerHTML")
        soup = BeautifulSoup(html, "html.parser")

        # 1° ciclo: deflagga i 4 default se è la prima iterazione
        if idx == 0:
            for dv in default_ids:
                try:
                    cb = wait.until(EC.element_to_be_clickable((
                        By.CSS_SELECTOR, f"#suppFileGds_more input[data-value_id='{dv}']"
                    )))
                    if cb.is_selected():
                        cb.click()
                        logging.debug(f"[Supplementary file] Deselezionato default: {dv}")
                except Exception as e:
                    logging.warning(f"[Supplementary file] Errore nel deselezionare {dv}: {e}")

        # Inserisce il nuovo valore via autocomplete
        try:
            inp = wait.until(EC.visibility_of_element_located((
                By.CSS_SELECTOR, "div.facets_dialog input.of_sel_inp"
            )))
            driver.execute_script("arguments[0].value = '';", inp)
            #time.sleep(0.3)
            inp.send_keys(val)
            time.sleep(0.2)
            inp.send_keys(Keys.RETURN)
            logging.debug(f"[Supplementary file] Inserito {val}")
            added_ids.append(val)
            #time.sleep(0.3)
        except Exception as e:
            logging.warning(f"[Supplementary file] Errore durante inserimento {val}: {e}")

    # Seleziona tutte le checkbox corrispondenti
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.facets_dialog")))
    customize = wait.until(EC.element_to_be_clickable((
        By.XPATH, "//a[@href='#suppFileGds_more' and contains(.,'Customize')]"
    )))
    ActionChains(driver).move_to_element(customize).click().perform()
    time.sleep(0.3)

    popup = wait.until(EC.visibility_of_element_located((
        By.CSS_SELECTOR, "#suppFileGds_more .facets_dialog"
    )))
    for val in added_ids:
        try:
            cb = wait.until(EC.element_to_be_clickable((
                By.CSS_SELECTOR, f"#suppFileGds_more input[data-value_id='{val}']"
            )))
            if not cb.is_selected():
                cb.click()
                logging.debug(f"[Supplementary file] Selezionato {val}")
        except Exception as e:
            logging.warning(f"[Supplementary file] Impossibile selezionare {val}: {e}")
        time.sleep(0.3)

    # Click finale su Show
    try:
        show_btn = wait.until(EC.element_to_be_clickable((By.ID, "suppFileGds_apply")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", show_btn)
        driver.execute_script("arguments[0].click();", show_btn)
        logging.debug("[Supplementary file] Show")
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "#suppFileGds_more .facets_dialog")))
    except Exception as e:
        logging.warning(f"[Supplementary file] Show non cliccabile: {e}")

    # Clicca nella sidebar
    group_label = "Supplementary file"
    for val in added_ids:
        try:
            sidebar_ul = driver.find_element(
                By.XPATH,
                f"//li[contains(@class,'filter_grp') and .//h3[normalize-space()='{group_label}']]//ul"
            )
            link = sidebar_ul.find_element(
                By.CSS_SELECTOR,
                f"a[data-value_id='{val}']"
            )
            logging.debug(f"[Supplementary file] {val} ")
            ActionChains(driver).move_to_element(link).click().perform()
            time.sleep(0.3)
        except Exception as e:
            logging.warning(f"[Supplementary file]  {val} : {e}")

    logging.info("Finished [Supplementary file]")

def apply_attribute_name_filter(driver, labels):
    time.sleep(0.5)
    logging.info("[Attribute name]")
    wait = WebDriverWait(driver, 4)
    labels = [v.strip().lower() for v in labels if v.strip()]
    if not labels:
        logging.info("No attribute names provided.")
        return

    default_ids = ["tissueGds", "strainGds"]

    added_ids = []

    # Inserimento valori via autocomplete
    for val in labels:
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.facets_dialog")))

        # Apri Customize...
        customize = wait.until(EC.element_to_be_clickable((
            By.XPATH, "//a[@href='#attNameGds_more' and contains(.,'Customize')]"
        )))
        ActionChains(driver).move_to_element(customize).click().perform()
        logging.debug(f"[Attribute name] Customize for: {val}")

        # Inserisci valore
        try:
            inp = wait.until(EC.visibility_of_element_located((
                By.CSS_SELECTOR, "div.facets_dialog input.of_sel_inp"
            )))
            driver.execute_script("arguments[0].value = '';", inp)
            inp.send_keys(val)
            time.sleep(0.2)
            inp.send_keys(Keys.RETURN)

            # Aspetta che compaia nella sidebar
            wait.until(EC.presence_of_element_located((
                By.CSS_SELECTOR, f"a[data-value_id='{val}']"
            )))

            added_ids.append(val)
            logging.debug(f"[Attribute name] Inserted: {val}")
        except Exception as e:
            logging.warning(f"[Attribute name] Autocomplete error for '{val}': {e}")

    # Riapri popup per la selezione checkbox
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.facets_dialog")))
    customize = wait.until(EC.element_to_be_clickable((
        By.XPATH, "//a[@href='#attNameGds_more' and contains(.,'Customize')]"
    )))
    ActionChains(driver).move_to_element(customize).click().perform()

    popup = wait.until(EC.visibility_of_element_located((
        By.CSS_SELECTOR, "#attNameGds_more .facets_dialog"
    )))

    # Deflagga i checkbox predefiniti (tissue, strain)
    for dv in default_ids:
        try:
            cb = wait.until(EC.element_to_be_clickable((
                By.CSS_SELECTOR, f"#attNameGds_more input[data-value_id='{dv}']"
            )))
            if cb.is_selected():
                cb.click()
                logging.debug(f"[Attribute name] Deselected default: {dv}")
        except Exception as e:
            logging.warning(f"[Attribute name] Failed to deselect {dv}: {e}")

    # Seleziona checkbox corrispondenti ai valori aggiunti
    for val in added_ids:
        try:
            cb = wait.until(EC.element_to_be_clickable((
                By.CSS_SELECTOR, f"#attNameGds_more input[data-value_id='{val}']"
            )))
            if not cb.is_selected():
                cb.click()
                logging.debug(f"[Attribute name] Selected {val}")
        except Exception as e:
            logging.warning(f"[Attribute name] Cannot select {val}: {e}")

    # Dentro apply_attribute_name_filter()
    try:
        # Click su "Show"
        show_btn = wait.until(EC.element_to_be_clickable((By.ID, "attNameGds_apply")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", show_btn)
        driver.execute_script("arguments[0].click();", show_btn)
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "#attNameGds_more .facets_dialog")))
        logging.debug("[Attribute name] Show clicked")
        time.sleep(0.3) 

    except UnexpectedAlertPresentException:
        alert = driver.switch_to.alert
        logging.warning(f"[Attribute name] Alert after Show: {alert.text}")
        alert.accept()


    # Clicca nella sidebar
    group_label = "Attribute name"
    for val in added_ids:
        try:
            sidebar_ul = driver.find_element(
                By.XPATH,
                f"//li[contains(@class,'filter_grp') and .//h3[normalize-space()='{group_label}']]//ul"
            )
            link = sidebar_ul.find_element(By.CSS_SELECTOR, f"a[data-value_id='{val}']")
            ActionChains(driver).move_to_element(link).click().perform()
            time.sleep(0.3)
            logging.debug(f"[Attribute name] Clicked `{val}` in sidebar")
        except Exception as e:
            logging.warning(f"[Attribute name] Sidebar `{val}`: {e}")


def apply_all_filters(driver, filters):

    _expand_filters(driver)

    if filters.get("organism"):
        autocomplete_filter(driver, "organismGds", filters["organism"], "organismGds_apply")

    if filters.get("study_type"):
        apply_study_type_filter(driver, filters["study_type"])

    if filters.get("subset_type"):
        click_filter_links(driver, "Subset variable type", filters["subset_type"])

    if filters.get("supp_file"):
        apply_supplementary_file_filter(driver, filters["supp_file"])
    if filters.get("attribute_name"):
        apply_attribute_name_filter(driver, filters["attribute_name"])
    if filters.get("date_range"):
        date_range = filters.get("date_range")

        if (
            isinstance(date_range, list) and
            len(date_range) == 2 and
            all(isinstance(d, list) and len(d) == 3 and all(isinstance(x, int) for x in d) for d in date_range)
        ):
            start_date, end_date = date_range
            logging.info(f"📆 date range: {start_date} -> {end_date}")
            apply_publication_date_range(driver, start_date, end_date)
        else:
            logging.warning("⚠️ Skipping date range filter: invalid or missing format")
    
################################################################

def get_resource_path(relative_path):
    """Ottiene il percorso assoluto per le risorse, funziona sia in development che quando impacchettato"""
    if getattr(sys, 'frozen', False):
        # Se l'applicazione è "frozen" (impacchettata con PyInstaller)
        application_path = sys._MEIPASS
    else:
        # Se siamo in modalità development
        application_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(application_path, relative_path)


def fetch_mesh_terms(pmid, max_retries=1):
    for attempt in range(max_retries):
        try:
            handle = Entrez.efetch(db="pubmed", id=pmid, rettype="medline", retmode="text")
            records = Medline.parse(handle)
            mesh_terms = []
            for record in records:
                if "MH" in record:
                    cleaned_terms = [term.strip() for term in record["MH"]]
                    mesh_terms.extend(cleaned_terms)
            handle.close()
            
            if not mesh_terms:
                return "No MeSH terms found"
            else:
                return mesh_terms
        
        except Exception as e:
            if attempt < max_retries - 1:  # se non è l'ultimo tentativo
                wait_time = (2 ** attempt) + random.uniform(0, 1) - random.uniform(0.005,0.010)  # backoff esponenziale con jitter
                logging.warning(f"Error retrieving MeSH terms for PMID {pmid}. Retrying in {wait_time:.2f} seconds.")
                logging.error(f"Details error: {str(e)}")
                time.sleep(wait_time)
            else:
                logging.error("Failed to retrieve MeSH terms for PMID {pmid} after {max_retries} attempts.")
                return f"Error fetching MeSH terms for PMID {pmid}: {str(e)}"

    return f"Failed to fetch MeSH terms for PMID {pmid} after {max_retries} attempts"
def extract_summary(page_source):
    try:
        soup = BeautifulSoup(page_source, 'html.parser')
        # Trova il <td> contenente il testo "Summary"
        summary_row = soup.find('td', string='Summary')
        if summary_row:
            # Trova il prossimo <td> con il testo della Summary
            summary_text = summary_row.find_next('td').get_text(strip=True)
            return summary_text
        else:
            return None
    except Exception as e:
        logging.error(f"Error extracting Summary: {e}")
        return None

def fetch_abstract(pmid):
  time.sleep(0.10)
  handle = Entrez.efetch(db="pubmed", id=pmid, rettype="Medline", retmode="text")
  rec = handle.read()
  handle.close()
  
  rec_file = StringIO(rec)
  medline_rec = Medline.read(rec_file)
  
  if "AB" in medline_rec:
    return medline_rec["AB"]
  else:
    return None

def search_total_pages_and_series_count(query, driver, filters=None):
    try:
        encoded_query = urllib.parse.quote(query)
        url = f"https://www.ncbi.nlm.nih.gov/gds/?term={encoded_query}"
        driver.get(url)
        
        logging.info(f"✅ Navigated directly to search results using advanced query")
        time.sleep(1) 
        # Try multiple selectors as GEO structure can vary
        try:
            series_count_element = driver.find_element(By.CSS_SELECTOR, "a[data-value_id='seriesGds'] + span")
            series_count_text = series_count_element.text
            series_count = int(re.sub(r'[(),]', '', series_count_text))
        except Exception:
            # Fallback: try reading from the filter list itself if the top tab is obscure
            try:
                # Based on your HTML: <li class="fil_val"><a href="#" data-value_id="seriesGds">Series</a><span class="fcount">(6)</span></li>
                series_count_element = driver.find_element(By.CSS_SELECTOR, "a[data-value_id='seriesGds'] ~ span.fcount")
                series_count_text = series_count_element.text
                series_count = int(re.sub(r'[(),]', '', series_count_text))
            except Exception:
                logging.warning("Could not determine series count, defaulting to 0")
                series_count = 0

        # 2. Get Total Pages
        # If results are few (e.g., 6), pagination controls (#pageno2) might NOT exist.
        try:
            total_pages_element = driver.find_element(By.CSS_SELECTOR, "input#pageno2")
            total_pages = int(total_pages_element.get_attribute("last"))
        except Exception:
            # If pagination input is missing, it usually means there is only 1 page.
            if series_count > 0:
                logging.info("Pagination input not found, assuming single page of results.")
                total_pages = 1
            else:
                total_pages = 0
        
        logging.info(f"Total pages found: {total_pages}")
        logging.info(f"Total series count: {series_count}")
        
        return total_pages, series_count
        
    except Exception as e:
        logging.error(f"❌ Failed to retrieve total pages or series count: {e}")
        # Return 1 page if we at least think we have results, to allow the scrape to proceed
        return 1, 0

def search_geo_datasets(query, driver, num_pages=3, max_retries=1, filters=None):
    gse_codes = set()
    
    encoded_query = urllib.parse.quote(query)
    
    url = f"https://www.ncbi.nlm.nih.gov/gds/?term={encoded_query}"
    logging.info(f"Navigating directly to URL: {url}")

    driver.get(url)

    page_html = driver.page_source.lower()
    if "500" in page_html and "internal server error" in page_html:
        logging.info("🔄 NCBI internal error detected, retrying...")

        retries = 0
        wait_time = 2  

        while retries < max_retries:
            driver.get(url)  
            page_html = driver.page_source

            soup = BeautifulSoup(page_html, "html.parser")
            login_button = soup.find("a", {"id": "account_login"})

            if login_button:
                logging.info("🔹 'Log in' button detected, waiting 1 second before continuing...")
                time.sleep(0.3)  

            if "500" not in page_html.lower() and "internal server error" not in page_html.lower():
                logging.info("✅ Successfully reloaded NCBI GEO DataSets page.")
                break  
            logging.warning(f"⚠️ Error 500 persists. Retrying in {wait_time} seconds...")
            time.sleep(wait_time)
            wait_time *= 2  
            retries += 1

        if retries == max_retries:
            logging.error("❌ Max retries reached. Could not bypass NCBI Internal Server Error.")
            return [] 

    for page in range(num_pages):
        try:
            WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.CLASS_NAME, "rprt"))
            )

            search_soup = BeautifulSoup(driver.page_source, "html.parser")
            dataset_titles = search_soup.find_all("div", class_="rprt")

            for title in dataset_titles:
                title_text = title.get_text().strip()
                gse_matches = re.findall(r"GSE\d+", title_text)
                for gse_code in gse_matches:
                    gse_codes.add(gse_code)

            next_buttons = driver.find_elements(By.CSS_SELECTOR, "a.next")

            if next_buttons:
                if page == num_pages - 1:
                    logging.info("⏳ Final page reached, waiting briefly...")
                    time.sleep(0.5)
                
                first_gse = None
                if dataset_titles:
                    first_gse_match = re.search(r"GSE\d+", dataset_titles[0].get_text())
                    if first_gse_match:
                        first_gse = first_gse_match.group()

                # Clicca Next
                next_buttons[0].click()
                logging.info(f"➡️ Clicked 'Next' to go to page {page + 2}")
                time.sleep(0.5)  

                # Attendi che il primo GSE della pagina precedente sparisca
                WebDriverWait(driver, 4).until(lambda d: (
                    first_gse not in d.page_source if first_gse else True
                ))

            else:
                logging.info(f"🔹 No 'Next' button found - Page {page + 1}")
                break

        except Exception as e:
            logging.error(f"❌ Error in search - Page {page + 1}: {e}")

            # Se c'è un errore nella paginazione, proviamo un semplice refresh per recuperare
            try:
                logging.info("Attempting to refresh the page to recover...")
                driver.refresh()
                time.sleep(2)
            except Exception as inner_e:
                logging.error(f"❌ Failed to restart/refresh search: {inner_e}")
                break  

    return list(gse_codes)

def extract_samples(driver):
    try:
        samples_table_xpath = "//td[starts-with(text(), 'Samples')]/following-sibling::td//table"
        sample_table = WebDriverWait(driver, 4).until(
            EC.presence_of_element_located((By.XPATH, samples_table_xpath))
        )

        sample_rows_xpath = ".//tr"
        sample_rows = sample_table.find_elements(By.XPATH, sample_rows_xpath)

        if not sample_rows:
            logging.warning("Sample rows not found.")
            return "Not found"
        
        samples = []
        for row in sample_rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) > 1:
                sample_id = cells[0].text.strip()  
                sample_description = cells[1].text.strip()  
                samples.append(sample_id)  

        return "; ".join(samples)  
    except TimeoutException:
        logging.warning("Timeout: Sample rows not found.")
        return "Not found"
    except Exception as e:
        logging.error(f'Error extracting Samples with links: {e}')
        return "Not found"
    
def extract_organisms(page_source):
    try:
        soup = BeautifulSoup(page_source, 'html.parser')
        
        organisms_row = soup.find('td', string=lambda x: x and ('Organisms' in x or 'Organism' in x))
        if organisms_row:
            organisms_cell = organisms_row.find_next('td')
            if organisms_cell:
                links = organisms_cell.find_all('a')
                organism_texts = []
                
                for link in links:
                    organism_name = link.text.strip()
                    organism_texts.append(organism_name)
                
                return ', '.join(organism_texts)
        
        return "Not found"
    except Exception as e:
        logging.error(f'Errore extracting Organisms: {e}')
        return "Not found"
def extract_platform(page_source):
    try:
        soup = BeautifulSoup(page_source, 'html.parser')

        platform_row = soup.find('td', string=lambda x: x and ('Platforms' in x or "Platform" in x))

        if platform_row:
            table = platform_row.find_next('table')
            if table:
                rows = table.find_all('tr')
                platform_data = []

                for row in rows:

                    gpl_cell = row.find('a')  
                    description_cell = row.find_all('td')[1]  

                    if gpl_cell and description_cell:
                        gpl_code = gpl_cell.text.strip()
                        description = description_cell.text.strip()
                        platform_data.append((gpl_code, description))  

                return platform_data  

        return "Not found"
    
    except Exception as e:
        logging.error(f'Error extracting Platform: {e}')
        return "Not found"
    
def extract_sample_count(driver):
    try:
        samples_td_xpath = '//td[contains(text(), "Samples")]'
        samples_td_element = WebDriverWait(driver, 4).until(
            EC.presence_of_element_located((By.XPATH, samples_td_xpath))
        )
        samples_td_text = samples_td_element.text.strip()
        sample_count = re.search(r'\((\d+)\)', samples_td_text).group(1)
        return int(sample_count)
    except TimeoutException:
        logging.error('Timeout: Unable to find Sample Count')
        return "Not found"
    except Exception as e:
        logging.error(f'Error extracting Sample Count {e}')
        return "Not found"

def extract_date(status): #accetta status per estrarre data
    match = re.search(r'(\w+) (\d{2}), (\d{4})', status) 
    if match:
        month_str = match.group(1) 
        day = match.group(2) 
        year = match.group(3) 
        months = {"Jan" :"01",
            "January": "01",
            "February": "02","Feb":"02",
            "March": "03", "Mar": "03",
            "April": "04", "Apr" : "04",
            "May": "05",
            "June": "06", "Jun" : "06",
            "July": "07", "Jul" : "07",
            "August": "08", "Aug" : "08",
            "September": "09", "Sep" :"09", "Sept" : "09",
            "October": "10", "Oct" :"10",
            "November": "11", "Nov" :"11",
            "December": "12" , "Dec" : "12"
        }
        month = months[month_str] 
        return f"{day}/{month}/{year}"
    return None

def create_output_file(dataframe, query, file_type):
    random_number = random.randint(1, 10000)
    random_letter = random.choice(string.ascii_letters)
    
    if file_type.lower() == 'excel':
        file_name = f"{query}_analysis_{random_number}{random_letter}.xlsx"

        dataframe.to_excel(file_name, index=False)
        
        workbook = load_workbook(file_name)
        sheet = workbook.active
        
        link_columns = [
            "Series Matrix Link", 
            "SOFT formatted family file(s) Link", 
            "MINiML formatted family file(s) Link",
            "Title/PMID", "Geo2R", "BioProject link", "Other link and GDV","SRA Run Selector","GSE", "Samples_1","Samples_2","Samples_3","Instrument_1","Instrument_2","Instrument_3"
        ]
        
        for column_name in link_columns:
            if column_name in dataframe.columns:
                link_col_idx = dataframe.columns.get_loc(column_name) + 1  
                
                for row_idx in range(2, len(dataframe) + 2):  
                    cell = sheet.cell(row=row_idx, column=link_col_idx)
                    cell_value = cell.value
                    
                    if isinstance(cell_value, str) and cell_value.startswith("https:"):
                        url_part= cell_value.split(" / ")[0]
                        cell.hyperlink = url_part 
                        cell.font = Font(color="0000FF", underline="single")  
        
        workbook.save(file_name)
        logging.info(f"Excel file saved as {file_name}")
    
    elif file_type.lower() == 'csv':
        file_name = f"{query}_analysis_{random_number}{random_letter}.csv"
        dataframe.to_csv(file_name, index=False)
        logging.info(f"CSV file saved as {file_name}")
    else:
        raise ValueError("File type must be either 'excel' or 'csv'")
    
    return file_name

def extract_title(page_source):
    try:
        soup = BeautifulSoup(page_source, 'html.parser')
        title_row = soup.find('td', string='Title')  
        if title_row:
            title_cell = title_row.find_next_sibling('td')  
            if title_cell:
                title_text = title_cell.get_text(strip=True)  
                return title_text
        return "Not found"
    except Exception as e:
        logging.error(f"Error extracting title: {e}")
        return "Not found"
    
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def extract_geo2r(page_source, gse):
    try:
        soup = BeautifulSoup(page_source, 'html.parser')

        geo2r_span = soup.find('span', id='geo2r')
        if geo2r_span:
            button = geo2r_span.find('button', id='geo2r_button')
            if button and 'Analyze with GEO2R' in button.get_text(strip=True):
                geo2r_link = f"https://www.ncbi.nlm.nih.gov/geo/geo2r/?acc={gse}"
                return geo2r_link
        return "Not found"
    except Exception as e:
        logging.error(f"Error extracting Geo2R for GSE {gse}: {e}")
        return "Not found"
    
def extract_bioproject(page_source):
    try:
        soup = BeautifulSoup(page_source, 'html.parser')
        
        bioproject_row = soup.find('td', string=lambda x: x and 'BioProject' in x)
        if bioproject_row:
            bioproject_cell = bioproject_row.find_next_sibling('td')
            if bioproject_cell:
                bioproject_link = bioproject_cell.find('a')  
                if bioproject_link:
                    bioproject_url = bioproject_link['href'].strip()  
                    
                    if bioproject_url.startswith("/bioproject"):
                        bioproject_url = f"https://www.ncbi.nlm.nih.gov{bioproject_url}"
                    
                    return bioproject_url
        
        bioproject_span = soup.find('span', class_='gp_id')
        if bioproject_span:
            bioproject_link = bioproject_span.find('a')  
            if bioproject_link:
                bioproject_url = bioproject_link['href'].strip()
                
                if bioproject_url.startswith("/bioproject"):
                    bioproject_url = f"https://www.ncbi.nlm.nih.gov{bioproject_url}"
                
                return bioproject_url

        return "Not found"
    except Exception as e:
        logging.error(f"Error extracting BioProject: {e}")
        return "Not found"

    
def extract_download_and_gdv(page_source, gse):
    try:
        soup = BeautifulSoup(page_source, 'html.parser')
        links = []

        gdv_button = soup.find('button', id='gdv_button')
        if gdv_button and 'See on Genome Data Viewer' in gdv_button.get_text(strip=True):
            gdv_link = f"https://www.ncbi.nlm.nih.gov/gdv/browser/?context=GEO&acc={gse}"
            links.append(gdv_link)
        
        download_button = soup.find('button', id='download_button')
        if download_button and 'Download RNA-seq counts' in download_button.get_text(strip=True):
            download_link = f"https://www.ncbi.nlm.nih.gov/geo/download/?acc={gse}"
            links.append(download_link)

        if links:
            return " _ ".join(links)
        
        return "Not found"
    except Exception as e:
        logging.error(f"Error extracting Download & GDV for GSE {gse}: {e}")
        return "Not found"
    
def extract_pmid_bs4(page_source):
    try:
        soup = BeautifulSoup(page_source, 'html.parser')
        citation_td = soup.find('td', string='Citation(s)')
        if citation_td:
            span = citation_td.find_next_sibling('td').find('span', class_='pubmed_id')
            if span:
                a_tag = span.find('a')
                if a_tag:
                    return a_tag.text.strip()
        return None
    except Exception as e:
        logging.error(f"BS4 PMID extraction failed: {e}")
        return None

def extract_sra_run_selector_link(page_source):
    soup = BeautifulSoup(page_source, "html.parser")
    link_tag = soup.find("a", string="SRA Run Selector")
    if link_tag and link_tag.get("href"):
        return f"https://www.ncbi.nlm.nih.gov{link_tag['href']}"
    return "Not found"

def minmax_normalize(values):
    valid_values = [v for v in values if v is not None and not np.isnan(v)]
    
    if not valid_values:
        return [0.0] * len(values)
        
    # Applichiamo max(0, v) per impedire che la Cosine Similarity negativa alteri la scala
    valid_values = [max(0, v) for v in valid_values]
    
    # Ancoriamo il minimo a 0
    vmin = 0 
    vmax = max(valid_values)
    
    if vmax == vmin:
        return [0.0] * len(values) # Se il max è 0, tutto è 0
    
    final_scores = []
    for v in values:
        if v is None or np.isnan(v):
            final_scores.append(0.0) # I valori nulli restano 0
        else:
            # Tagliamo a 0 i valori originali prima di normalizzarli
            v_clipped = max(0, v)
            final_scores.append((v_clipped - vmin) / (vmax - vmin))
            
    return final_scores


def calculate_textual_scores(df: pd.DataFrame, search_keywords: list, target_mesh_list: list, original_query: str = "") -> pd.DataFrame:
    logging.info("--- Starting Advanced Textual Score Calculation (Qdrant + Champion Logic) ---")
    
    if df.empty:
        logging.warning("DataFrame is empty, skipping score calculation.")
        return df
    bias_phrase = "this superseries is composed of the subseries listed below"
    
    # Rimuoviamo il bias dal Summary: se è presente, azzeriamo il testo
    df['Summary'] = df['Summary'].fillna('').astype(str).apply(
        lambda x: "" if bias_phrase in x.lower() else x
    )
    # Pulizia preliminare del testo
    df['Clean_Title'] = df['Title/PMID'].astype(str).apply(
        lambda x: x.split(' / ', 1)[1] if ' / ' in x else ('' if x.strip().startswith('http') else x)
    )
    # Uniamo Titolo e Summary per dare più contesto al modello
    df['Full_Text'] = (df['Clean_Title'] + " " + df['Summary'].fillna('')).str.lower()

    # ---------------------------------------------------------
    # 1. CALCOLO K_SCORE (Keyword Match - Champion Logic)
    # ---------------------------------------------------------
    if not search_keywords:
        df['K_score'] = 0.0
    else:
        valid_keywords = [k.lower().strip() for k in search_keywords if k.strip()]
        if not valid_keywords:
             df['K_score'] = 0.0
        else:
            df['K_count'] = 0.0
            for kw in valid_keywords:
                df['K_count'] += df['Full_Text'].apply(lambda x: 1.0 if kw in x else 0.0)
            
            max_found_in_batch = df['K_count'].max()
            logging.info(f"🏆 Champion Paper found: {max_found_in_batch} keywords out of {len(valid_keywords)} requested.")
            
            if max_found_in_batch == 0:
                df['K_score'] = 0.0
            else:
                ratio = df['K_count'] / max_found_in_batch
                df['K_score'] = ratio.pow(1.35)
            
            df.drop(columns=['K_count'], inplace=True)

    # ---------------------------------------------------------
    # 2. CALCOLO S_SCORE (Semantic Similarity con QDRANT)
    # ---------------------------------------------------------
    try:
        collection_name = "geo_papers"
        
        # A. Gestione Collezione
        if qdrant.collection_exists(collection_name):
            qdrant.delete_collection(collection_name)
        
        qdrant.create_collection(
            collection_name=collection_name,
            vectors_config=VectorParams(size=384, distance=Distance.COSINE),
        )

        # B. Calcola Embeddings
        logging.info("Embedding")
        documents_text = df['Full_Text'].tolist()
        doc_embeddings = model.encode(documents_text)
        df['Vector'] = [json.dumps(vec) for vec in doc_embeddings.tolist()]
        points = []
        for idx, embedding in enumerate(doc_embeddings):
            vector_list = embedding.tolist()
            points.append(PointStruct(
                id=idx,
                vector=vector_list,
                payload={"text": documents_text[idx]} 
            ))
        
        # C. Caricamento in Qdrant
        logging.info(f"{len(points)} vectors are being upserted into Qdrant...")
        qdrant.upsert(collection_name=collection_name, points=points)

        # D. Embedding della Query
        # Combiniamo la query GEO originale con le keyword per dare al modello
        # il contesto completo dell'intenzione di ricerca del ricercatore.
        # Esempio: original_query="breast cancer" + keywords="rna tumor tp53"
        # → query_text="breast cancer rna tumor tp53" (molto più ricco di solo "rna tumor tp53")
        keyword_text = " ".join(search_keywords)
        query_parts = [p for p in [original_query.strip(), keyword_text.strip()] if p]
        query_text = " ".join(query_parts) if query_parts else "biomedical genomics"
        logging.info(f"Semantic query text: '{query_text}'")
        query_vector = model.encode(query_text).tolist()

        # E. Cerca (Search) -> USA query_points
        logging.info("Semantic Search in Qdrant...")
        
        # --- MODIFICA CRUCIALE QUI ---
        search_result_obj = qdrant.query_points(
            collection_name=collection_name,
            query=query_vector, # Si chiama 'query', non 'query_vector' in questa versione
            limit=len(df) 
        )
        # L'oggetto ritornato ha un attributo .points che contiene la lista dei risultati
        search_hits = search_result_obj.points

        # F. Mapping dei risultati
        s_scores_map = {hit.id: hit.score for hit in search_hits}
        
        # Assegna score
        df['S_raw'] = [s_scores_map.get(i, 0.0) for i in range(len(df))]

        # G. Normalizza
        df['S_score'] = minmax_normalize(df['S_raw'].tolist())
        
        logging.info("✅ Qdrant semantic scoring completed successfully.")

    except Exception as e:
        logging.error(f"❌ Qdrant calculation failed: {e}")
        # Debug di sicurezza
        try:
            logging.error(f"Errore specifico: {type(e).__name__}, Args: {e.args}")
        except:
            pass
        df['S_score'] = 0.0

    # ---------------------------------------------------------
    # 3. CALCOLO M_SCORE (MeSH Terms) - Invariato
    # ---------------------------------------------------------
    clean_targets = [t.strip() for t in target_mesh_list if t.strip()]
    valid_mesh_cols = [col for col in clean_targets if col in df.columns]
    m_weight_active = False 
    
    if not valid_mesh_cols:
        df['M_score'] = 0.0
        logging.info("ℹ️ No matching MeSH columns found in scraper data.")
    else:
        matches_sum = df[valid_mesh_cols].sum(axis=1)
        if matches_sum.sum() == 0:
             df['M_score'] = 0.0
        else:
             df['M_score'] = matches_sum.apply(lambda x: 1.0 if x >= 1 else 0.0)
             m_weight_active = True

    # ---------------------------------------------------------
    # 4. CALCOLO FINALE T_SCORE (Pesi Dinamici)
    # ---------------------------------------------------------
    w_s = 0.50  
    w_k = 0.30  
    w_m = 0.20  
    
    if not m_weight_active:
        logging.info("⚠️ MeSH score inactive. Redistributing weights.")
        w_s = 0.60
        w_k = 0.40
        w_m = 0.00

    df['T_score'] = (df['S_score'] * w_s) + (df['K_score'] * w_k) + (df['M_score'] * w_m)
    
    df.drop(columns=['Clean_Title', 'Full_Text', 'S_raw'], inplace=True, errors='ignore')

    return df

def calculate_bibliographic_scores(df: pd.DataFrame) -> pd.DataFrame:

    def parse_date(date_str):
        try:
            return datetime.strptime(date_str, '%d/%m/%Y').date()
        except Exception:
            return None
            
    df['Pub_Date'] = df['Date'].apply(parse_date)
    valid_dates = df['Pub_Date'].dropna()
    
    if not valid_dates.empty:
        # Usiamo i giorni rispetto alla data più vecchia nel BATCH (non da anno 0).
        # Con toordinal() + vmin=0, una data del 2020 vale ~737059 e una del 2024 vale ~738521:
        # entrambe si normalizzano a >0.998 — differenza trascurabile e inutile.
        # Con min-del-batch come riferimento, lo spread reale (es. 4 anni = ~1460 giorni)
        # diventa l'intera scala 0→1, rendendo il punteggio significativo.
        batch_min_ordinal = valid_dates.apply(lambda x: x.toordinal()).min()
        df['R_raw'] = df['Pub_Date'].apply(
            lambda x: (x.toordinal() - batch_min_ordinal) if x else None
        )
        df['R_score'] = minmax_normalize(df['R_raw'].tolist())
    else:
        df['R_score'] = 0.5 

    df['Citations_Count'] = df['Citations_Count'].fillna(0).astype(float)
    
    df['C_log'] = df['Citations_Count'].apply(log1p)
    
    df['C_score'] = minmax_normalize(df['C_log'].tolist())
    
    df['B_score'] = (df['R_score'] * 0.5) + (df['C_score'] * 0.5)
    
    return df

def fetch_citation_count(pmid):
    try:
        if not pmid:
            return 0

        handle = Entrez.elink(
            dbfrom="pubmed", 
            id=str(pmid), 
            linkname="pubmed_pubmed_citedin"
        )
        record = Entrez.read(handle)
        handle.close()

        linksetdb = record[0].get("LinkSetDb", [])

        # Se non è una lista → nessuna citazione
        if not isinstance(linksetdb, list) or len(linksetdb) == 0:
            logging.info(f"No citations found for PMID {pmid}")
            return 0

        citation_count = len(linksetdb[0].get("Link", []))
        logging.info(f"Citations found for PMID {pmid}: {citation_count}")
        return citation_count

    except Exception as e:
        logging.error(f"Error fetching citations for PMID {pmid}: {e}")
        return 0

def extract_PMID_tile(pmid):
    try:
        handle = Entrez.efetch(db="pubmed", id=str(pmid), rettype="Medline", retmode="text")
        record = handle.read()
        handle.close()

        medline_rec = Medline.read(StringIO(record))
        title1 = medline_rec.get("TI", "Not found")
        return title1
    except Exception as e:
        logging.error(f"Error fetching title for PMID {pmid}: {e}")
        return "Not found"

def extract_study_type(page_source):
    try:
        soup = BeautifulSoup(page_source, 'html.parser')

        # Cerca esattamente la riga che contiene "Experiment type"
        for row in soup.find_all('tr'):
            cells = row.find_all('td')
            if len(cells) >= 2:
                header = cells[0].get_text(strip=True)

                if re.search(r'^experiment\s*type$', header, re.IGNORECASE):
                    study_type = cells[1].get_text(strip=True)
                    logging.info(f"Study type extracted: {study_type}")
                    return study_type

    except Exception as e:
        logging.error(f"Error extracting study type: {str(e)}")
    
    return None
    
def process_data(query, email, num_pages1=2, keyword1="tumor,bladder", m_s="Red,Brown", file_type="excel", mode="normal", generate_file=True,remove=True,filters=None):
    print(f"query: {query}")
    print(f"email: {email}")
    print(f"num_pages1: {num_pages1}")
    print(f"keyword1: {keyword1}")
    print(f"m_s: {m_s}")
    print(f"file_type: {file_type}")
    print(f"mode: {mode}")
    print(f"generate_file: {generate_file}")
    print(f"remove: {remove}")

    # Controlliamo se keyword1 è None
    if keyword1 is None:
        print("ERRORE: keyword1 è None!")
        keyword1 = ""

    # Controlliamo se m_s è None
    if m_s is None:
        print("ERRORE: m_s è None!")
        m_s = ""

    gse_counter = 0
    logging.info(f"Starting processing for the query: {query}")
    #=ChromeDriverManager().install()
    try:
        # Configura Chrome options
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--log-level=3')
        options.add_argument('--incognito')
        #driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
        # Aggiungi ulteriori opzioni per la stabilità
        options.add_argument('--start-maximized')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        #options.add_argument('--disable-gpu')
        #options.add_experimental_option('excludeSwitches', ['enable-logging'])
        prefs = {"profile.default_content_setting_values.geolocation": 2,"profile.default_content_setting_values.notifications": 2}
        options.add_experimental_option("prefs", prefs)

        driver = webdriver.Chrome(options=options)
        logging.info("WebDriver initialized successfully; no geolocation or notification will be used")
        
    except Exception as e:
        logging.error(f"Error initializing WebDriver: {str(e)}")
        raise Exception(f"Error initializing Chrome: {str(e)}")
    
    columns = ["Title/PMID", "GSE", "Date", "Instrument_1","Instrument_2","Instrument_3", "Platform", "Organisms", "Samples_Count", "Samples_1","Samples_2","Samples_3", "Summary", "Series Matrix Link", "SOFT formatted family file(s) Link","MINiML formatted family file(s) Link","BioProject link","Geo2R", "Other link and GDV","SRA Run Selector",'Citations_Count', 'Study_Type_Extracted']
    Entrez.email = email
    keywords = []
    M_S = []
    
    user_keywords = keyword1.split(",")
    for i in range(len(user_keywords)):
        user_keywords[i] = user_keywords[i].lower()
    duplicates = [kw for kw in user_keywords if kw in keywords]
    if duplicates:
        logging.info(f"Duplicated keyword founded: {', '.join(duplicates)}")
    else: 
        logging.info("No duplicated keywords")
    
    keywords.extend(user_keywords)
    columns.extend(user_keywords)
    
    user_MS = [MS.strip() for MS in m_s.split(',')]
    columns.extend(user_MS)
    
    gse_codes = search_geo_datasets(query, driver, num_pages1,filters=filters)  # avvia la funzione
    logging.info(f"GSE codes founded: {gse_codes}")
    logging.info(f"Total GSE: {len(gse_codes)}")
    logging.info("_________________________________________________________________________________________________")

    df = pd.DataFrame(columns=columns)  # crea un dataframe con quelle colonne

    # Lista per memorizzare i PMDI trovati
    pmdi_list = []

    # Copia della lista GSE per iterare ed eventualmente rimuovere elementi
    gse_codes_copy = gse_codes.copy()
    total_gse = len(gse_codes_copy)
    
    pubmed_base = "https://pubmed.ncbi.nlm.nih.gov/"
    gse_counter = 0
    
    for gse in gse_codes_copy:
        logging.info(f"Processing GSE: {gse}")
        gse_counter += 1
        
        logging.info(f"Processing GSE: {gse_counter}")
        url = f'https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc={gse}' 
        driver.get(url)  # accedi all'url
        
        try:
    # Verifica comunque che l'elemento in fondo sia presente
            WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.ID, 'vdp'))
            )
            time.sleep(0.1)
        except TimeoutException:
            try:
                logging.info("Due to a slow internet connection or high site traffic, the long wait function has been activated to prevent data loss.")
                time.sleep(0.5)
            except TimeoutException:
                continue
        
        html = driver.page_source  # estraiamo codice html
        pattern_id = r'id="([^"]+)"'  # regex (espressione regolare,pattern) per trovare ID
        id_trovati = re.findall(pattern_id, html)  # trova tutte le corrispondenze con quel pattern_id nel codice html della pagina
        
        # crea nuova lista espressione per ogni elemento in iterabile se condizione vera
        my_id = [string for string in id_trovati if string.isdigit()]  # stringa per ogni elemento all'interno degli ID se questo è un numero, aggiungi alla lista
    
        found_pmdi = False  # inizializza variabile, falsa all'inizio 
        pmdi_already_checked = False
        pmdi1 = None
        for i in my_id:
            logging.info(f"GSE {gse_counter} of {total_gse} - Processing for GSE: {gse}")
            try:
                time.sleep(0.01)

                if not pmdi_already_checked:
                    pmdi = extract_pmid_bs4(driver.page_source)
                    pmdi_already_checked = True

                    if pmdi:
                        pmdi1 = pmdi
                        pmdi_list.append(pmdi1)
                        found_pmdi = True
                        logging.info(f"PMID found in Citation box: {pmdi1}")
                        link_pubmed = pubmed_base + pmdi1
                        logging.info(f"Link to Pubmed: {link_pubmed}")
                    else:
                        logging.info(" No PMID found via BS4")

                if not found_pmdi:
                    continue  # Salta il blocco successivo se non trovato
                
                # Estrai la data dalla sezione "Status"
                try:
                    status_xpath = '//td[normalize-space(text())="Status"]/following-sibling::td'
                    status_element = WebDriverWait(driver, 4).until(
                        EC.presence_of_element_located((By.XPATH, status_xpath))
                    )  # aspettiamo finchè non troviamo
                    status_text = status_element.text  # trasformiamolo in testo
                    logging.info(f'Status found: {status_text}')
                    
                    # Format the extracted status text
                    status_text = extract_date(status_text)  # applichiamo la funzione delle date
                    
                except Exception as e:
                    status_text = "Not found"
                    logging.error(f'Errore extracting status for {gse}: {e}')
                    
                # Dopo aver caricato la pagina e assicurato che sia completamente caricata
                page_source = driver.page_source
                try:
                    platform_text = extract_platform(page_source)
                    if platform_text != "Not found":
                        gpl_links = [f"https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc={gpl}" for gpl, _ in platform_text]
                        gpl_codes = ";".join(gpl_links)
                        platform_descriptions = ", ".join([desc for _, desc in platform_text])
                        logging.info(f"GPL Codes : {gpl_codes} , Platform description : {platform_descriptions}")
                    else:
                        gpl_codes = "Not found"
                        platform_descriptions = "Not found"
                except Exception:
                    gpl_codes = "Not found"
                    platform_descriptions = "Not found"
                    logging.error(f"Error extracting platform for {gse}; {e}")

                try:
                    # Ottieni il contenuto HTML della pagina
                    page_source = driver.page_source
                    # Estrai gli organismi usando la funzione BeautifulSoup
                    organism_text = extract_organisms(page_source)
                    logging.info(f"Organism {organism_text} found")
                except Exception as e:
                    organism_text = "Not found"
                    logging.error(f'Error extracting organism for {gse}: {e}')
                
                try:
                    samples_text = extract_samples(driver)
                    if samples_text != "Not found":
                        base_url = "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc="
                        samples_links = "; ".join([f"{base_url}{gsm}" for gsm in samples_text.split("; ")])
                    else:
                        samples_links = "Not found"
                except Exception as e:
                    samples_text = "Not found"
                    logging.error(f'Error extracting samples for {gse}: {e}')
                
                try:
                    count = extract_sample_count(driver)
                    logging.info(f"Sample count {count} for {gse}")
                except Exception as e:
                    count = "Not found"
                    logging.error(f'Error extracting Sample count for {gse}: {e}')

                try:
                    summary_text = extract_summary(page_source)
                    if summary_text:
                        logging.info(f"Summary extracted for GSE {gse}: {summary_text}")
                    else:
                        logging.info(f"No Summary found for GSE {gse}")
                except Exception as e:
                    summary_text = "NaN"  # Assicurati di assegnare NaN se fallisce
                    logging.error(f"Error extracting Summary for GSE {gse}: {e}")

                try:
                    study_type = extract_study_type(page_source)
                    logging.info(f"Study type for {gse}: {study_type}")
                except Exception as e:
                    study_type = "Not found"
                    logging.error(f"Error extracting study type for {gse}: {e}")

                citation_count = 0
                if found_pmdi and pmdi1:
                    citation_count = fetch_citation_count(pmdi1)
                title_pmid= extract_PMID_tile(pmdi1) if found_pmdi and pmdi1 else "Not found"
                pmid_link=f"https://pubmed.ncbi.nlm.nih.gov/{pmdi1} / {title_pmid}"
                series_matrix_link= f"https://ftp.ncbi.nlm.nih.gov/geo/series/GSE{gse[3:6]}nnn/{gse}/matrix/"
                SOFT_link= f"https://ftp.ncbi.nlm.nih.gov/geo/series/GSE{gse[3:6]}nnn/{gse}/soft/"
                mini_link = f"https://ftp.ncbi.nlm.nih.gov/geo/series/GSE{gse[3:6]}nnn/{gse}/miniml/"
                try:
                    page_source = driver.page_source
                    geo2r_link = extract_geo2r(page_source, gse)
                    if geo2r_link != "Not found":
                        logging.info(f"GEO2R link for GSE {gse}: {geo2r_link}")
                    else:
                        logging.info(f"Geo2R Not found in {gse}")
                except Exception as e:
                    geo2r_link = "Not found"
                    logging.error(f"Error extracting GEO2R for GSE {gse}: {e}")
                try:
                    page_source = driver.page_source
                    sra_selector_link = extract_sra_run_selector_link(page_source)
                    if sra_selector_link != "Not found":
                        logging.info(f"SRA Run Selector link for GSE {gse}: {sra_selector_link}")
                    else:
                        logging.info(f"SRA Run Selector NOT found for {gse}")
                except Exception as e:
                    sra_selector_link = "Not found"
                    logging.error(f"Error extracting SRA Run Selector for GSE {gse}: {e}")
                try:
                    page_source = driver.page_source
                    bioproject_link = extract_bioproject(page_source)
                    if bioproject_link != "Not found":
                        logging.info(f"Bioproject link {bioproject_link} found")
                    else:
                        continue
                except Exception as e:
                    logging.info(f"error on bioproject for {gse}")
                try:
                    download_and_gdv_link = extract_download_and_gdv(page_source, gse)
                except Exception as e:
                    download_and_gdv_link = "Not found"

                samples_list = samples_links.split('; ') if samples_links != "Not found" else []
                instruments_list = gpl_codes.split(';') if gpl_codes != "Not found" else []
                new_row = {'Title/PMID': pmid_link, 'GSE': url, 'Date': status_text, "Platform": platform_descriptions, "Organisms": organism_text, "Samples_Count": count, "Summary": summary_text, "Series Matrix Link": series_matrix_link, "SOFT formatted family file(s) Link": SOFT_link, "MINiML formatted family file(s) Link": mini_link, "BioProject link": bioproject_link, "Geo2R":geo2r_link, "Other link and GDV": download_and_gdv_link,"SRA Run Selector":sra_selector_link, "Citations_Count": citation_count, "Study_Type_Extracted": study_type}  # aggiungi alla colonna  queste cose
                for i in range(3):  # Supponiamo un massimo di 3 campioni (Sample_1, Sample_2, Sample_3)
                    col_name = f'Samples_{i+1}'
                    new_row[col_name] = samples_list[i] if i < len(samples_list) else 0
                for i in range(3):  # Supponiamo un massimo di 3 strumenti (Instrument_1, Instrument_2, Instrument_3)
                    col_name = f'Instrument_{i+1}'
                    new_row[col_name] = instruments_list[i].strip() if i < len(instruments_list) else 0
                for col in columns[22:]:  # a partire dalla sesta colonna di columns
                    new_row[col] = 0  # Inizializza con 0 la riga corrispondente alla colonna
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)  # concatena a df un nuovo dataframe contenente la nuova riga

                abstract = fetch_abstract(pmdi1)
                abstract = abstract.lower()
                if abstract:
                    for keyword in keywords:  # scorri le keyword
                        if keyword in abstract:  # se trovi la keyword
                            df.loc[df['Title/PMID'] == pmid_link, keyword] = 1
                            logging.info(f'Keyword "{keyword}" found in abstract for PMID{pmdi1}')
                
                mesh_terms = fetch_mesh_terms(pmdi1, max_retries=1)
                if isinstance(mesh_terms, str) and mesh_terms.startswith("Error"):
                    logging.warning(f"Unable to retrieve MeSH terms for PMID {pmdi1}: {mesh_terms}")
                else:
                    #logging.info(f"MeSH terms successfully retrieved for PMID {pmdi1}")
                    logging.info(f"MeSH terms for PMID {pmdi1}: {mesh_terms}")

                    for mesh_term in mesh_terms:
                        if mesh_term in df.columns:
                            matching_rows = df.loc[df['Title/PMID'].str.contains(pmdi1)]
                            if not matching_rows.empty:
                                df.loc[df['Title/PMID'].str.contains(pmdi1), mesh_term] = 1
                            else:
                                logging.warning(f"No row found")

                    logging.info("_________________________________________________________________________________________________")


            except Exception as e:
                logging.error(f'Error with ID {i} for {gse}: {e}')
        
        if not found_pmdi and mode=="normal":
            logging.warning(f"No PMID found for: {gse} ")
            gse_codes.remove(gse)
            logging.info("_________________________________________________________________________________________________")  # se non abbiamo trovato nessun id pmdi


################### ULTRA MODE ###################
        elif not found_pmdi and mode == "ultra":
            try:
                logging.info("No PMID found - GSE Analysis in ultra mode")
                
                # Estrazione e controllo immediato del Titolo (FAIL FAST)
                page_source = driver.page_source
                title_text = extract_title(page_source)
                
                if not title_text or title_text == "Not found":
                    logging.warning(f"⚠️ Title not found for {gse}. Skipping entire analysis to save time.")
                    logging.info("_________________________________________________________________________________________________")
                    continue # Salta subito al prossimo ciclo for
                
                logging.info(f"Title extracted for GSE {gse}: {title_text}")

                # Estrazione Status/Date con timeout ridotto
                try:
                    status_xpath = '//td[normalize-space(text())="Status"]/following-sibling::td'
                    status_element = WebDriverWait(driver, 4).until(
                        EC.presence_of_element_located((By.XPATH, status_xpath))
                    )
                    status_text = extract_date(status_element.text.strip())
                except Exception:
                    status_text = "Not found"

                # Estrazione Platform
                try:
                    platform_text = extract_platform(page_source)
                    if platform_text != "Not found":
                        gpl_links = [f"https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc={gpl}" for gpl, _ in platform_text]
                        gpl_codes = ";".join(gpl_links)
                        platform_descriptions = ", ".join([desc for _, desc in platform_text])
                        logging.info(f"GPL Codes : {gpl_codes} , Platform description : {platform_descriptions}")
                    else:
                        gpl_codes = "Not found"
                        platform_descriptions = "Not found"
                except Exception as e:
                    gpl_codes = "Not found"
                    platform_descriptions = "Not found"
                    logging.error(f"Error extracting platform for {gse}: {e}")

                # Estrazione Organism
                try:
                    organism_text = extract_organisms(page_source)
                    logging.info(f"Organism {organism_text} found")
                except Exception as e:
                    organism_text = "Not found"
                    logging.error(f'Error extracting organism for {gse}: {e}')

                # Estrazione Samples
                try:
                    samples_text = extract_samples(driver)
                    if samples_text != "Not found":
                        base_url = "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc="
                        samples_links = "; ".join([f"{base_url}{gsm}" for gsm in samples_text.split("; ")])
                    else:
                        samples_links = "Not found"
                except Exception as e:
                    samples_text = "Not found"
                    logging.error(f'Error extracting samples for {gse}: {e}')

                # Estrazione Sample Count
                try:
                    count = extract_sample_count(driver)
                    logging.info(f"Sample count {count} for {gse}")
                except Exception as e:
                    count = "Not found"
                    logging.error(f'Error extracting Sample count for {gse}: {e}')

                # Estrazione Summary
                try:
                    summary_text = extract_summary(page_source)
                    if summary_text:
                        logging.info(f"Summary extracted for GSE {gse}: {summary_text}")
                    else:
                        logging.info(f"No Summary found for GSE {gse}")
                except Exception as e:
                    summary_text = "NaN"
                    logging.error(f"Error extracting Summary for GSE {gse}: {e}")

                # Estrazione Study Type
                try:
                    study_type = extract_study_type(page_source)
                    logging.info(f"Study type for {gse}: {study_type}")
                except Exception as e:
                    study_type = "Not found"
                    logging.error(f"Error extracting study type for {gse}: {e}")

                # Estrazione Links vari
                citation_count = 0 
                # (Nota: qui citation_count rimane 0 perché found_pmdi è False in questo blocco)

                try:
                    geo2r_link = extract_geo2r(page_source, gse)
                    if geo2r_link != "Not found":
                        logging.info(f"Geo2R link found for {gse}")
                    else:
                        logging.info(f"GEO2R link NOT found for {gse}")
                except Exception as e:
                    geo2r_link = "Not found"
                    logging.error(f"Error extracting GEO2R for GSE {gse}")

                series_matrix_link = f"https://ftp.ncbi.nlm.nih.gov/geo/series/GSE{gse[3:6]}nnn/{gse}/matrix/"
                soft_link = f"https://ftp.ncbi.nlm.nih.gov/geo/series/GSE{gse[3:6]}nnn/{gse}/soft/"
                mini_link = f"https://ftp.ncbi.nlm.nih.gov/geo/series/GSE{gse[3:6]}nnn/{gse}/miniml/"
                
                try:
                    sra_selector_link = extract_sra_run_selector_link(page_source)
                    if sra_selector_link != "Not found":
                        logging.info(f"SRA Run Selector link for GSE {gse}: {sra_selector_link}")
                    else:
                        logging.info(f"SRA Run Selector NOT found for {gse}")
                except Exception as e:
                    sra_selector_link = "Not found"
                    logging.error(f"Error extracting SRA Run Selector for GSE {gse}: {e}")

                try:
                    bioproject_link = extract_bioproject(page_source)
                    if bioproject_link != "Not found":
                        logging.info(f"Bioproject link {bioproject_link} found")
                    else:
                        # Qui c'era un 'continue' rischioso nel tuo codice originale. 
                        # Meglio settarlo a "Not found" e proseguire per non perdere il GSE.
                        bioproject_link = "Not found"
                except Exception as e:
                    bioproject_link = "Not found"
                    logging.info(f"error on bioproject for {gse}")

                # Titolo Finale (Ultra Mode usa il titolo estratto dalla pagina)
                title_pmid_value = title_text 

                try:
                    download_and_gdv_link = extract_download_and_gdv(page_source, gse)
                    logging.info(f"Download & GDV link for GSE {gse}: {download_and_gdv_link}")
                except Exception as e:
                    download_and_gdv_link = "Not found"
                    logging.error(f"Error extracting Download & GDV for GSE {gse}: {e}")

                # --- CREAZIONE RIGA DATAFRAME ---
                samples_list = samples_links.split('; ') if samples_links != "Not found" else []
                instruments_list = gpl_codes.split(';') if gpl_codes != "Not found" else []
                
                new_row = {
                    'Title/PMID': title_pmid_value, 
                    'GSE': url, 
                    'Date': status_text, 
                    "Platform": platform_descriptions, 
                    "Organisms": organism_text, 
                    "Samples_Count": count, 
                    "Summary": summary_text, 
                    "Series Matrix Link": series_matrix_link, 
                    "SOFT formatted family file(s) Link": soft_link, 
                    "MINiML formatted family file(s) Link": mini_link, 
                    "BioProject link": bioproject_link, 
                    "Geo2R": geo2r_link, 
                    "Other link and GDV": download_and_gdv_link,
                    "SRA Run Selector": sra_selector_link, 
                    "Citations_Count": citation_count, 
                    "Study_Type_Extracted": study_type
                }

                # Gestione dinamica Samples/Instruments (fino a 3)
                for i in range(3):
                    col_name = f'Samples_{i+1}'
                    new_row[col_name] = samples_list[i] if i < len(samples_list) else 0
                for i in range(3):
                    col_name = f'Instrument_{i+1}'
                    new_row[col_name] = instruments_list[i].strip() if i < len(instruments_list) else 0
                
                # Riempimento zeri per keyword/mesh (colonne dalla 22 in poi)
                # Attenzione: columns[22:] deve essere coerente con come hai definito 'columns' all'inizio
                for col in columns[22:]:
                    new_row[col] = 0
                
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

                # --- RICERCA KEYWORD NEL SUMMARY (Ultra Mode non ha abstract PubMed) ---
                if summary_text and summary_text != "NaN":
                    summary_text_lower = summary_text.lower()
                    keyword_found = False
                    for keyword in keywords:
                        if keyword in summary_text_lower:
                            # Nota: usiamo title_text come chiave per trovare la riga
                            df.loc[df['Title/PMID'] == title_text, keyword] = 1
                            logging.info(f'Keyword "{keyword}" found in Summary for Title "{title_text}"')
                            keyword_found = True
                    if not keyword_found:
                        logging.info(f"No keywords found in Summary for Title : {title_text}")
                
                logging.info("_________________________________________________________________________________________________")

            except Exception as e:
                logging.error(f"Error processing Title for GSE {gse}: {e}")

    #df['sum_score'] = df.iloc[:, 7:].sum(axis=1)
    #print('List of found PMID:', pmdi_list)
    #print('List of linked GSE:', gse_codes)
    #pd.set_option('display.max_columns', None)
    #print('DataFrame:', df)
    logging.info("Starting relevance score calculation...")

    # 1. Prepara liste input
    search_keywords_list = [k.strip().lower() for k in keyword1.split(',') if k.strip()]
    
    if m_s:
        target_mesh_list = [x.strip() for x in m_s.split(',')]
    else:
        target_mesh_list = []

    # 2. Chiama direttamente la NUOVA funzione textual
    df = calculate_textual_scores(df, search_keywords_list, target_mesh_list, original_query=query)

    # 3. Chiama la funzione bibliografica
    df = calculate_bibliographic_scores(df)

    # 4. Calcola il Relevance finale (sovrascrivendo la logica vecchia)
    DEFAULT_T = 0.75
    DEFAULT_B = 0.25
    
    # Rinomina per coerenza
    df.rename(columns={'T_score': 'T', 'B_score': 'B'}, inplace=True)
    
    # Formula finale
    df['Relevance'] = (df['T'] * DEFAULT_T) + (df['B'] * DEFAULT_B)
    df['Relevance_Display'] = df['Relevance'].round(3).astype(str)
    
    logging.info("Relevance scores calculated successfully")
    
    # ============ FINE CALCOLO SCORE ============


    score_columns = [
        'K_score', 'S_raw', 'S_score', 'M_score',
        'R_raw', 'R_score', 'C_log', 'C_score',
        'T_score', 'B_score', 'Relevance', 'missing_S','T','B','Pub_Date'
    ]
    existing_score_cols = [c for c in score_columns if c in df.columns]

    key_col = 'GSE' if 'GSE' in df.columns else (df.columns[1] if len(df.columns) > 1 else None)
    if key_col:
        df_scores = df.loc[:, [key_col] + existing_score_cols].copy()
    else:
        df_scores = df.loc[:, existing_score_cols].copy()

    logging.info(f"df_scores prepared with columns: {df_scores.columns.tolist()}")
    df_scores_dict=df_scores.to_dict(orient="records")

    df.drop(columns=existing_score_cols, errors='ignore', inplace=True)

    if remove:
        df = df.loc[:, (df != 0).any(axis=0)]
    # Creare il file di output e ottenere il percorso completo e il nome del file
    if generate_file:
        f= create_output_file(df, query, file_type)
        file_path=f
        df_scores.to_csv("df_scores.csv", index=False)
    else:
        file_path= "No file generated"
    df.replace({np.nan: None}, inplace=True)
    json_data=df.to_json(orient="records")
    df_dict= df.to_dict(orient="records")
    #print("file created:",f)
    return {
        "df_scores": df_scores_dict,
        "pmid_list": pmdi_list,
        "gse_codes": gse_codes,
        "dataframe": df_dict,
        "file_path": file_path,
        "json_data": json_data
    }
if __name__ == "__main__":
    query = "breast cancer"
    email= ""
    num_pages1 = 3
    keyword1 = "breast,tumor,immunotherapy,chemotherapy,rna,dna,protein"
    m_s = "Human, Mice"
    file_type = "excel"
    mode="ultra"
    generate_file=True
    remove=False
    result = process_data(query, email, num_pages1, keyword1, m_s, file_type,mode,generate_file,remove)
    print(result)